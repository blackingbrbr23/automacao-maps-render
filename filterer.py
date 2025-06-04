# -------------------------------------------------------------------
# filterer.py
# -------------------------------------------------------------------
# Captura informações de cada perfil de links.txt e salva em Excel.
# -------------------------------------------------------------------

import os
import sys
import time
import logging
import re
import requests
from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side

# -------------------------------------------------------------
# Configuração de logging
# -------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

# -------------------------------------------------------------
# URL do painel remoto (vai buscar em variável de ambiente)
# -------------------------------------------------------------
PAINEL_URL = os.environ.get("PAINEL_URL", "https://painel-api-k2v2.onrender.com/command")

def get_mac():
    """
    Retorna o endereço MAC principal do sistema (igual a scanner.get_mac).
    """
    try:
        if sys.platform.startswith("win"):
            import subprocess
            output = subprocess.check_output("getmac", encoding="cp1252", errors="ignore")
            macs = re.findall(r"([0-9A-Fa-f]{2}[-:]){5}[0-9A-Fa-f]{2}", output)
            return macs[0].replace("-", ":").lower() if macs else None
        else:
            import subprocess
            output = subprocess.check_output("ifconfig", encoding="utf-8", errors="ignore")
            macs = re.findall(r"([0-9A-Fa-f]{2}[:]){5}[0-9A-Fa-f]{2}", output)
            return macs[0].lower() if macs else None
    except Exception as e:
        logger.error(f"Erro ao obter MAC: {e}")
        return None

def verificar_liberacao(mac, tentativas=3, intervalo=2) -> bool:
    """
    Mesmo comportamento que em scanner.verificar_liberacao.
    """
    for tentativa in range(1, tentativas + 1):
        try:
            ip = requests.get("https://api.ipify.org", timeout=5).text
            r = requests.get(PAINEL_URL, params={"mac": mac, "public_ip": ip}, timeout=10)
            r.raise_for_status()
            return r.json().get("ativo", False)
        except Exception as e:
            logger.error(f"Tentativa {tentativa}/{tentativas} falhou: {e}")
            if tentativa < tentativas:
                time.sleep(intervalo)
    return False

def salvar_infor(driver, informacoes_path):
    """
    Extrai dados do perfil atual (já aberto no driver) e salva em Excel:
      - Nome
      - Telefone
      - Estrelas
      - Avaliações
      - Endereço
      - Nicho (botões de categoria)
      - Data da última atualização
      - Link
    """
    try:
        wait = WebDriverWait(driver, 10)
        elemento_data = wait.until(EC.presence_of_element_located((By.XPATH,
            "//div[contains(@class, 'lchoPb') or contains(@class', 'mqX5ad')]")))
        data_atualizacao = elemento_data.text.strip()
        link_atual = driver.current_url

        driver.find_element(By.CLASS_NAME, "fKm1Mb").click()
        time.sleep(4)

        numero_texto = wait.until(EC.presence_of_element_located(
            (By.XPATH, '//button[@class="CsEnBe" and contains(@aria-label, "Telefone:")]'))
        ).find_element(By.CLASS_NAME, "Io6YTe").text.strip()

        estrelas_texto = driver.find_element(By.CSS_SELECTOR, "div.fontDisplayLarge").text
        avaliacoes = wait.until(EC.visibility_of_element_located(
            (By.XPATH, "//span[contains(text(),'avaliações')]")
        )).text

        h1 = wait.until(EC.visibility_of_element_located((By.TAG_NAME, "h1")))
        texto_completo = h1.text
        nome = texto_completo.split("–")[0].strip() if "–" in texto_completo else texto_completo

        botoes = driver.find_elements(By.CSS_SELECTOR, 'button.DkEaL')
        nicho = ", ".join([b.text for b in botoes if b.text.strip()])
        endereco_texto = driver.find_element(By.CSS_SELECTOR, 'div.Io6YTe').text

        # Definir caminho do Excel em dados/
        os.makedirs(os.path.dirname(informacoes_path), exist_ok=True)
        if os.path.exists(informacoes_path):
            wb = load_workbook(informacoes_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append([
                "Nome", "Telefone", "Estrelas", "Avaliações", "Endereço",
                "Nicho", "Data Última Atualização", "Link"
            ])

        telefones_existentes = {
            ws.cell(row=r, column=2).value
            for r in range(2, ws.max_row + 1)
        }
        if numero_texto in telefones_existentes:
            logger.warning(f"⚠️ Registro com telefone {numero_texto} já existe.")
            return

        ws.append([
            nome, numero_texto, estrelas_texto, avaliacoes,
            endereco_texto, nicho, data_atualizacao, link_atual
        ])

        thin = Side(border_style="thin", color="CCCCCC")
        borda = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                cell.border = borda
                wrap = False if cell.column == ws.max_column else True
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap)

        for col in ws.columns:
            mx = max((len(str(c.value)) for c in col if c.value), default=0)
            ws.column_dimensions[col[0].column_letter].width = mx * 1.2 + 2

        wb.save(informacoes_path)
        logger.info("✅ Planilha formatada e salva com sucesso!")
    except Exception as e:
        logger.error(f"Erro ao salvar info: {e}")

def vericar_street_view(driver):
    """
    Se encontrar elemento com texto “Google Street View”, considera que não deve salvar.
    Retorna False (pulando) ou True (prosseguir).
    """
    try:
        elem = driver.find_element(By.CLASS_NAME, "ilzTS")
        if "Google Street View" in elem.text:
            logger.info("✅ Perfil criado pelo Google. Pulando.")
            return False
        return True
    except:
        return True

def data_check(driver, informacoes_path):
    """
    Pega a data de última atualização do perfil (texto tipo “Atualizado em abr. de 2025”).
    Se for mais antiga que 3 meses a partir de hoje, chama salvar_infor.
    """
    try:
        wait = WebDriverWait(driver, 10)
        texto = wait.until(EC.presence_of_element_located((By.XPATH,
            "//div[contains(@class, 'lchoPb') or contains(@class, 'mqX5ad')]"))).text.strip()

        from datetime import datetime, date
        from dateutil.relativedelta import relativedelta

        agora = datetime.now()
        meses = {
            'jan.': 1, 'fev.': 2, 'mar.': 3, 'abr.': 4,
            'mai.': 5, 'jun.': 6, 'jul.': 7, 'ago.': 8,
            'set.': 9, 'out.': 10, 'nov.': 11, 'dez.': 12
        }
        part = texto.split(" - ")
        data_str = part[1].lower() if len(part) > 1 else texto.lower()
        m_nome, _, a_str = data_str.partition(" de ")
        mes = meses.get(m_nome.strip())
        ano = int(a_str.strip()) if a_str else agora.year
        data_perfil = date(ano, mes, 1)
        data_limite = date(agora.year, agora.month, 1) - relativedelta(months=3)

        if data_perfil >= data_limite:
            logger.info("✅ Perfil atualizado nos últimos 3 meses.")
        else:
            logger.warning("⚠️ Perfil desatualizado há mais de 3 meses. Salvando...")
            salvar_infor(driver, informacoes_path)
    except Exception as e:
        logger.error(f"Erro data_check: {e}")
        if vericar_street_view(driver):
            salvar_infor(driver, informacoes_path)

def process_profiles(driver, informacoes_path):
    """
    Para cada link de perfil que já está aberto no driver,
    clica na imagem do container, chama data_check e possivelmente salva.
    """
    try:
        containers = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CLASS_NAME, "ofKBgf"))
        )
        logger.info(f"🔍 Encontrados {len(containers)} containers.")
        for cont in containers:
            try:
                img = cont.find_element(By.CLASS_NAME, "DaSXdd")
                driver.execute_script("arguments[0].scrollIntoView({block:'center'})", img)
                time.sleep(0.5)
                img.click()
                time.sleep(1)
                data_check(driver, informacoes_path)
            except Exception as e:
                logger.warning(f"Erro ao processar container: {e}")
                continue
    except Exception as e:
        logger.error(f"Nenhum container encontrado: {e}")

def run_filter(dados_dir):
    """
    Ponto de entrada para processamento:
      - Verifica liberação no painel remoto
      - Lê todos os arquivos .txt em dados/ (geralmente links.txt)
      - Para cada URL, abre no Selenium headless e chama process_profiles
      - Salva em dados/informacoes.xlsx
    """
    mac = get_mac()
    if not mac:
        logger.error("Não foi possível obter o endereço MAC. Abortando.")
        return

    logger.info("→ Verificando liberação junto ao painel remoto…")
    if not verificar_liberacao(mac):
        logger.error("Cliente BLOQUEADO. Abortando.")
        return

    LINKS_PATH = os.path.join(dados_dir, "links.txt")
    INFORMACOES_PATH = os.path.join(dados_dir, "informacoes.xlsx")

    if not os.path.exists(dados_dir):
        logger.error(f"Pasta 'dados' não encontrada em {dados_dir}")
        return

    # Lê todos os links de todos os .txt dentro de dados_dir
    links = []
    for arquivo in os.listdir(dados_dir):
        if arquivo.endswith(".txt"):
            caminho_arquivo = os.path.join(dados_dir, arquivo)
            with open(caminho_arquivo, 'r', encoding='utf-8') as f:
                novos_links = [l.strip() for l in f if l.strip()]
                logger.info(f"📄 {arquivo}: {len(novos_links)} links carregados.")
                links.extend(novos_links)

    if not links:
        logger.warning("Nenhum link encontrado nos arquivos da pasta 'dados'.")
        return

    # Configura Selenium Firefox em modo headless
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    service = Service(GeckoDriverManager().install())
    driver = webdriver.Firefox(service=service, options=options)

    for url in links:
        logger.info(f"🔗 Acessando: {url}")
        driver.get(url)
        time.sleep(5)
        process_profiles(driver, INFORMACOES_PATH)

    driver.quit()
    logger.info("🏁 Automação de filtragem concluída!")
